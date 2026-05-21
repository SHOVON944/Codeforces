
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, FormulaRule, CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime, date
import random

wb = Workbook()

# ══════════════════════════════════════════════════════════════
# COLORS
# ══════════════════════════════════════════════════════════════
# Dark theme palette
BG_DARK       = "0D1117"
BG_CARD       = "161B22"
BG_MID        = "1C2333"
CYAN          = "00BFFF"
CYAN_DARK     = "008FBF"
GREEN         = "00FF88"
GREEN_DARK    = "00AA55"
GREEN_MID     = "00CC66"
YELLOW        = "FFD700"
YELLOW_DARK   = "FFA500"
RED           = "FF4444"
RED_DARK      = "CC0000"
PURPLE        = "BD93F9"
PINK          = "FF79C6"
ORANGE        = "FFB86C"
WHITE         = "F8F8F2"
GRAY_L        = "8B949E"
GRAY_D        = "30363D"
GOLD          = "F1C40F"
TEAL          = "1ABC9C"
BLUE          = "3498DB"

def clr(hex_code):
    return hex_code.upper()

def fill(hex_code):
    return PatternFill("solid", fgColor=clr(hex_code))

def font(color=WHITE, bold=False, size=11, name="Calibri", italic=False):
    return Font(name=name, size=size, bold=bold, italic=italic, color=clr(color))

def align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border(colors=None):
    c = clr(GRAY_D) if not colors else clr(colors)
    s = Side(style="thin", color=c)
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    s = Side(style="medium", color=clr(CYAN))
    return Border(left=s, right=s, top=s, bottom=s)

def style_header(cell, bg=BG_MID, fg=CYAN, sz=12, bold=True):
    cell.fill = fill(bg)
    cell.font = font(fg, bold, sz)
    cell.alignment = align()
    cell.border = thin_border(GRAY_D)

def style_cell(cell, bg=BG_DARK, fg=WHITE, sz=10, bold=False, h="center"):
    cell.fill = fill(bg)
    cell.font = font(fg, bold, sz)
    cell.alignment = align(h)
    cell.border = thin_border(GRAY_D)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def merge_title(ws, cell_range, text, bg=BG_DARK, fg=CYAN, sz=16, bold=True):
    ws.merge_cells(cell_range)
    c = ws[cell_range.split(":")[0]]
    c.value = text
    c.fill = fill(bg)
    c.font = font(fg, bold, sz)
    c.alignment = align()
    c.border = thin_border(CYAN)

# ══════════════════════════════════════════════════════════════
# SHEET 1: DASHBOARD
# ══════════════════════════════════════════════════════════════
ws_dash = wb.active
ws_dash.title = "🏠 Dashboard"
ws_dash.sheet_view.showGridLines = False
ws_dash.sheet_properties.tabColor = clr(CYAN)

# Background entire sheet
for row in ws_dash.iter_rows(min_row=1, max_row=60, min_col=1, max_col=20):
    for cell in row:
        cell.fill = fill(BG_DARK)

# Title Banner
for r in range(1, 5):
    for c in range(1, 21):
        ws_dash.cell(r, c).fill = fill(BG_CARD)

ws_dash.merge_cells("A1:T1")
c = ws_dash["A1"]
c.value = "🎓  SHOVON — Personal Academic & Life Manager"
c.fill = fill(BG_CARD)
c.font = Font(name="Calibri", size=20, bold=True, color=clr(CYAN))
c.alignment = align()

ws_dash.merge_cells("A2:T2")
c = ws_dash["A2"]
c.value = "✦  University  ·  Attendance  ·  Study  ·  Tasks  ·  Finance  ·  CGPA  ✦"
c.fill = fill(BG_CARD)
c.font = font(GRAY_L, False, 11, italic=True)
c.alignment = align()

ws_dash.merge_cells("A3:T3")
c = ws_dash["A3"]
c.value = f"Last Updated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}"
c.fill = fill(BG_CARD)
c.font = font(YELLOW, False, 10)
c.alignment = align()

ws_dash.row_dimensions[4].height = 6

# KPI Cards Row 1
kpi_row = 5
kpi_data = [
    ("📚 Total Courses", "=COUNTA('📋 Courses'!B5:B30)", CYAN, "A"),
    ("🎯 Attendance %", "='📊 Attendance'!S5", GREEN, "D"),
    ("✅ Done Study %", "='📖 Study Tracker'!L5", YELLOW, "G"),
    ("💰 Total Expense", "=SUM('💸 Expenses'!E5:E200)", ORANGE, "J"),
    ("⭐ Current CGPA", "='🏆 CGPA'!H5", PURPLE, "M"),
    ("📅 Pending Tasks", "=COUNTIF('📝 Tasks'!F5:F200,\"Pending\")", RED, "P"),
]

kpi_cols = {"A": 1, "D": 4, "G": 7, "J": 10, "M": 13, "P": 16}

for (label, formula, color, col_letter) in kpi_data:
    c_num = kpi_cols[col_letter]
    end_col = get_column_letter(c_num + 2)
    ws_dash.merge_cells(f"{col_letter}{kpi_row}:{end_col}{kpi_row}")
    ws_dash.merge_cells(f"{col_letter}{kpi_row+1}:{end_col}{kpi_row+1}")
    ws_dash.merge_cells(f"{col_letter}{kpi_row+2}:{end_col}{kpi_row+2}")
    ws_dash.merge_cells(f"{col_letter}{kpi_row+3}:{end_col}{kpi_row+3}")

    c = ws_dash[f"{col_letter}{kpi_row}"]
    c.value = label; c.fill = fill(BG_CARD)
    c.font = font(color, True, 10); c.alignment = align()

    c2 = ws_dash[f"{col_letter}{kpi_row+1}"]
    c2.value = formula; c2.fill = fill(BG_CARD)
    c2.font = font(WHITE, True, 18); c2.alignment = align()

    c3 = ws_dash[f"{col_letter}{kpi_row+2}"]
    c3.fill = fill(color); c3.value = ""

    c4 = ws_dash[f"{col_letter}{kpi_row+3}"]
    c4.fill = fill(BG_DARK); c4.value = ""

ws_dash.row_dimensions[kpi_row].height = 20
ws_dash.row_dimensions[kpi_row+1].height = 36
ws_dash.row_dimensions[kpi_row+2].height = 4
ws_dash.row_dimensions[kpi_row+3].height = 8

# Section headers
def dash_section(ws, row, col_start, col_end, title, color):
    ws.merge_cells(f"{get_column_letter(col_start)}{row}:{get_column_letter(col_end)}{row}")
    c = ws.cell(row, col_start)
    c.value = title; c.fill = fill(BG_CARD)
    c.font = font(color, True, 12); c.alignment = align("left")
    c.border = Border(left=Side(style="thick", color=clr(color)))

row = 10
dash_section(ws_dash, row, 1, 19, "  📌  Quick Navigation & Module Status", CYAN)

nav_items = [
    ("📋 Courses", "📋 Courses", "Configure all courses & semester info", GREEN),
    ("📊 Attendance", "📊 Attendance", "Mark & track class attendance", CYAN),
    ("📖 Study Tracker", "📖 Study Tracker", "Monitor study progress per course", YELLOW),
    ("📝 Tasks", "📝 Tasks", "Daily task & routine manager", ORANGE),
    ("💸 Expenses", "💸 Expenses", "Daily buy & expense tracker", RED),
    ("💰 Money Manager", "💰 Money Manager", "Fund accounts & transactions", GREEN),
    ("🏆 CGPA", "🏆 CGPA", "Marks, GPA & CGPA calculation", PURPLE),
    ("📅 Calendar", "📅 Calendar", "Attendance calendar view", PINK),
    ("📑 Fee Manager", "📑 Fee Manager", "Semester registration fees", GOLD),
]

for i, (label, sheet, desc, color) in enumerate(nav_items):
    r = row + 2 + i
    ws_dash.row_dimensions[r].height = 22
    c1 = ws_dash.cell(r, 1)
    c1.value = f"  {label}"; c1.fill = fill(BG_CARD)
    c1.font = font(color, True, 11); c1.alignment = align("left")
    ws_dash.merge_cells(f"A{r}:E{r}")

    c2 = ws_dash.cell(r, 6)
    c2.value = desc; c2.fill = fill(BG_DARK)
    c2.font = font(GRAY_L, False, 10); c2.alignment = align("left")
    ws_dash.merge_cells(f"F{r}:S{r}")

ws_dash.row_dimensions[row + 1].height = 5

# Column widths for dashboard
for i in range(1, 21):
    set_col_width(ws_dash, i, 10)

# ══════════════════════════════════════════════════════════════
# SHEET 2: COURSES
# ══════════════════════════════════════════════════════════════
ws_c = wb.create_sheet("📋 Courses")
ws_c.sheet_view.showGridLines = False
ws_c.sheet_properties.tabColor = clr(GREEN)

for row in ws_c.iter_rows(min_row=1, max_row=80, min_col=1, max_col=20):
    for cell in row:
        cell.fill = fill(BG_DARK)

merge_title(ws_c, "A1:R1", "📋  COURSE & SEMESTER CONFIGURATION", BG_CARD, GREEN, 16)
ws_c.merge_cells("A2:R2")
ws_c["A2"].value = "  Configure all courses, credits, instructors and semester details"
ws_c["A2"].fill = fill(BG_CARD)
ws_c["A2"].font = font(GRAY_L, False, 10, italic=True)
ws_c["A2"].alignment = align("left")

# Semester selector
ws_c.merge_cells("A3:D3")
ws_c["A3"].value = "🔵 Active Semester:"
ws_c["A3"].fill = fill(BG_CARD)
ws_c["A3"].font = font(CYAN, True, 11)
ws_c["A3"].alignment = align("left")

ws_c["E3"].value = "1st Semester"
ws_c["E3"].fill = fill(BG_MID)
ws_c["E3"].font = font(YELLOW, True, 11)
ws_c["E3"].alignment = align()
ws_c["E3"].border = thin_border(CYAN)

dv_sem = DataValidation(type="list", formula1='"1st Semester,2nd Semester,3rd Semester,4th Semester,5th Semester,6th Semester,7th Semester,8th Semester"', allow_blank=False)
dv_sem.error = "Invalid semester selection"
dv_sem.errorTitle = "Input Error"
dv_sem.prompt = "Select the active semester"
dv_sem.promptTitle = "Semester"
ws_c.add_data_validation(dv_sem)
dv_sem.add("E3")

headers = ["#", "Semester", "Course Code", "Course Title", "Instructor", "Credit Hours",
           "Mid Marks", "Final Marks", "Assignment", "Lab", "Attendance Marks",
           "Total Marks", "Grade", "Status", "Start Date", "End Date", "Notes"]

h_colors = [GRAY_L, CYAN, GREEN, YELLOW, ORANGE, PURPLE, BLUE, BLUE, PINK, PINK, TEAL, GOLD, RED, GREEN, CYAN, CYAN, GRAY_L]

ws_c.row_dimensions[4].height = 28
for col, (h, hc) in enumerate(zip(headers, h_colors), 1):
    cell = ws_c.cell(4, col, h)
    cell.fill = fill(BG_MID)
    cell.font = font(hc, True, 10)
    cell.alignment = align()
    cell.border = thin_border(CYAN)

col_widths = [4, 14, 13, 30, 22, 12, 10, 10, 11, 8, 16, 12, 8, 12, 12, 12, 20]
for i, w in enumerate(col_widths, 1):
    set_col_width(ws_c, i, w)

# Sample course data
courses_data = [
    ("1st Semester", "CSE101", "Introduction to Programming", "Dr. Rahman", 3, 25, 40, 10, 0, 10),
    ("1st Semester", "MATH101", "Calculus I", "Prof. Khan", 3, 25, 40, 10, 0, 10),
    ("1st Semester", "PHY101", "Physics I", "Dr. Ahmed", 3, 25, 40, 10, 5, 10),
    ("1st Semester", "ENG101", "English Communication", "Ms. Fatema", 2, 25, 40, 10, 0, 10),
    ("2nd Semester", "CSE201", "Data Structures", "Dr. Islam", 3, 25, 40, 10, 0, 10),
    ("2nd Semester", "MATH201", "Calculus II", "Prof. Hossain", 3, 25, 40, 10, 0, 10),
]

row_bg = [BG_DARK, BG_CARD]
for i, (sem, code, title, teacher, credit, mid, final, asgn, lab, att) in enumerate(courses_data):
    r = 5 + i
    bg = row_bg[i % 2]
    ws_c.row_dimensions[r].height = 22
    vals = [i+1, sem, code, title, teacher, credit, mid, final, asgn, lab, att,
            f"=SUM(G{r}:K{r})", f'=IF(L{r}>=80,"A+",IF(L{r}>=75,"A",IF(L{r}>=70,"A-",IF(L{r}>=65,"B+",IF(L{r}>=60,"B",IF(L{r}>=55,"B-",IF(L{r}>=50,"C+",IF(L{r}>=45,"C",IF(L{r}>=40,"D","F")))))))))',
            "Active", "", "", ""]
    for col, val in enumerate(vals, 1):
        c = ws_c.cell(r, col, val)
        c.fill = fill(bg)
        c.font = font(WHITE, False, 10)
        c.alignment = align()
        c.border = thin_border(GRAY_D)

    ws_c.cell(r, 1).font = font(GRAY_L, False, 10)
    ws_c.cell(r, 2).font = font(CYAN, False, 10)
    ws_c.cell(r, 3).font = font(GREEN, True, 10)
    ws_c.cell(r, 4).font = font(YELLOW, False, 10)
    ws_c.cell(r, 5).font = font(ORANGE, False, 10)

# Status validation
dv_status = DataValidation(type="list", formula1='"Active,Completed,Dropped"')
ws_c.add_data_validation(dv_status)
dv_status.add(f"N5:N50")

# Semester validation in courses
dv_sem2 = DataValidation(type="list", formula1='"1st Semester,2nd Semester,3rd Semester,4th Semester,5th Semester,6th Semester,7th Semester,8th Semester"')
ws_c.add_data_validation(dv_sem2)
dv_sem2.add("B5:B50")

# Credit validation
dv_credit = DataValidation(type="whole", operator="between", formula1="1", formula2="6")
dv_credit.error = "Credit must be 1-6"
dv_credit.errorTitle = "Invalid Credit"
ws_c.add_data_validation(dv_credit)
dv_credit.add("F5:F50")

# Marks validation
dv_marks = DataValidation(type="decimal", operator="between", formula1="0", formula2="100")
dv_marks.error = "Marks must be 0-100"
ws_c.add_data_validation(dv_marks)
dv_marks.add("G5:K50")

# Conditional formatting - grade colors
ws_c.conditional_formatting.add("M5:M50",
    FormulaRule(formula=['M5="A+"'], fill=fill(GREEN_DARK), font=font(WHITE, True)))
ws_c.conditional_formatting.add("M5:M50",
    FormulaRule(formula=['M5="A"'], fill=fill(GREEN_MID), font=font(WHITE, True)))
ws_c.conditional_formatting.add("M5:M50",
    FormulaRule(formula=['M5="F"'], fill=fill(RED_DARK), font=font(WHITE, True)))

# Summary row
r_sum = 5 + len(courses_data) + 1
ws_c.merge_cells(f"A{r_sum}:K{r_sum}")
ws_c.cell(r_sum, 1).value = "TOTALS & AVERAGES"
ws_c.cell(r_sum, 1).fill = fill(BG_MID)
ws_c.cell(r_sum, 1).font = font(CYAN, True, 11)
ws_c.cell(r_sum, 1).alignment = align("right")
ws_c.cell(r_sum, 12).value = f"=AVERAGE(L5:L{r_sum-2})"
ws_c.cell(r_sum, 12).fill = fill(BG_MID)
ws_c.cell(r_sum, 12).font = font(GOLD, True, 11)
ws_c.cell(r_sum, 12).alignment = align()

# Also add CGPA summary box
ws_c.merge_cells("A{0}:D{0}".format(r_sum + 2))
ws_c.cell(r_sum+2, 1).value = "📌 Total Credits Registered:"
ws_c.cell(r_sum+2, 1).fill = fill(BG_CARD)
ws_c.cell(r_sum+2, 1).font = font(PURPLE, True, 11)
ws_c.cell(r_sum+2, 1).alignment = align("left")
ws_c.cell(r_sum+2, 5).value = f"=SUM(F5:F{r_sum-2})"
ws_c.cell(r_sum+2, 5).fill = fill(BG_CARD)
ws_c.cell(r_sum+2, 5).font = font(WHITE, True, 13)
ws_c.cell(r_sum+2, 5).alignment = align()

# ══════════════════════════════════════════════════════════════
# SHEET 3: ATTENDANCE
# ══════════════════════════════════════════════════════════════
ws_att = wb.create_sheet("📊 Attendance")
ws_att.sheet_view.showGridLines = False
ws_att.sheet_properties.tabColor = clr(CYAN)

for row in ws_att.iter_rows(min_row=1, max_row=220, min_col=1, max_col=22):
    for cell in row:
        cell.fill = fill(BG_DARK)

merge_title(ws_att, "A1:V1", "📊  ATTENDANCE TRACKER", BG_CARD, CYAN, 16)
ws_att["A2"].value = "  Mark attendance, track presence/absence, monitor attendance % per course"
ws_att.merge_cells("A2:V2")
ws_att["A2"].fill = fill(BG_CARD)
ws_att["A2"].font = font(GRAY_L, False, 10, italic=True)
ws_att["A2"].alignment = align("left")

# Summary stats row
ws_att.merge_cells("A3:C3"); ws_att["A3"].value = "📅 Filter Semester:"
ws_att["A3"].fill = fill(BG_CARD); ws_att["A3"].font = font(CYAN, True, 10); ws_att["A3"].alignment = align("left")
ws_att["D3"].value = "All Semesters"; ws_att["D3"].fill = fill(BG_MID)
ws_att["D3"].font = font(YELLOW, True, 10); ws_att["D3"].alignment = align()
ws_att["D3"].border = thin_border(CYAN)
dv_att_sem = DataValidation(type="list", formula1='"All Semesters,1st Semester,2nd Semester,3rd Semester,4th Semester,5th Semester,6th Semester,7th Semester,8th Semester"')
ws_att.add_data_validation(dv_att_sem); dv_att_sem.add("D3")

att_headers = ["#", "Date", "Day", "Semester", "Course Code", "Course Title", "Instructor",
               "Class Time", "Status", "Extra Class", "Online/Offline", "Topic Covered",
               "Signature Count", "Notes", "Week No", "Month", "Year",
               "Running Present", "Running Absent", "Running Total", "Att. %", "Remarks"]

att_colors = [GRAY_L, CYAN, PURPLE, BLUE, GREEN, YELLOW, ORANGE, TEAL, RED,
              PINK, CYAN, GRAY_L, GOLD, GRAY_L, PURPLE, CYAN, CYAN,
              GREEN, RED, WHITE, GOLD, ORANGE]

ws_att.row_dimensions[4].height = 28
for col, (h, hc) in enumerate(zip(att_headers, att_colors), 1):
    c = ws_att.cell(4, col, h)
    c.fill = fill(BG_MID)
    c.font = font(hc, True, 10)
    c.alignment = align()
    c.border = thin_border(CYAN)

att_widths = [4, 12, 8, 14, 12, 22, 18, 12, 10, 11, 13, 22, 14, 18, 8, 10, 6, 13, 13, 13, 8, 16]
for i, w in enumerate(att_widths, 1):
    set_col_width(ws_att, i, w)

days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
months = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
att_sample = [
    ("1st Semester", "CSE101", "Introduction to Programming", "Dr. Rahman", "08:00 AM - 09:30 AM", "Present", "No", "Offline", "Variables & Data Types", 2),
    ("1st Semester", "MATH101", "Calculus I", "Prof. Khan", "10:00 AM - 11:30 AM", "Present", "No", "Offline", "Limits", 2),
    ("1st Semester", "PHY101", "Physics I", "Dr. Ahmed", "12:00 PM - 01:30 PM", "Absent", "No", "Offline", "Motion", 0),
    ("1st Semester", "CSE101", "Introduction to Programming", "Dr. Rahman", "08:00 AM - 09:30 AM", "Present", "No", "Online", "Control Flow", 2),
    ("1st Semester", "MATH101", "Calculus I", "Prof. Khan", "10:00 AM - 11:30 AM", "Present", "Yes", "Offline", "Derivatives", 2),
    ("2nd Semester", "CSE201", "Data Structures", "Dr. Islam", "09:00 AM - 10:30 AM", "Present", "No", "Offline", "Arrays", 2),
    ("2nd Semester", "CSE201", "Data Structures", "Dr. Islam", "09:00 AM - 10:30 AM", "Present", "No", "Offline", "Linked Lists", 2),
    ("2nd Semester", "MATH201", "Calculus II", "Prof. Hossain", "11:00 AM - 12:30 PM", "Absent", "No", "Offline", "Integration", 0),
]

sample_dates = ["2026-01-05", "2026-01-05", "2026-01-05", "2026-01-07", "2026-01-07",
                "2026-02-02", "2026-02-04", "2026-02-04"]

for i, (row_data, d_str) in enumerate(zip(att_sample, sample_dates)):
    r = 5 + i
    sem, code, title, teacher, time, status, extra, mode, topic, sigs = row_data
    d = datetime.strptime(d_str, "%Y-%m-%d")
    bg = BG_DARK if i % 2 == 0 else BG_CARD

    vals = [i+1, d.strftime("%d-%m-%Y"), d.strftime("%a"), sem, code, title, teacher,
            time, status, extra, mode, topic, sigs, "",
            f"=WEEKNUM(B{r})", f"=TEXT(B{r},\"mmmm\")", f"=YEAR(B{r})",
            f'=COUNTIFS($I$5:I{r},"Present",$E$5:E{r},E{r})',
            f'=COUNTIFS($I$5:I{r},"Absent",$E$5:E{r},E{r})',
            f"=R{r}+S{r}",
            f'=IF(T{r}=0,0,R{r}/T{r})',
            f'=IF(U{r}>=0.75,"✅ Good",IF(U{r}>=0.60,"⚠️ Warning","❌ Low"))']

    for col, val in enumerate(vals, 1):
        c = ws_att.cell(r, col, val)
        c.fill = fill(bg)
        c.font = font(WHITE, False, 10)
        c.alignment = align()
        c.border = thin_border(GRAY_D)

    ws_att.row_dimensions[r].height = 20
    ws_att.cell(r, 2).font = font(CYAN, False, 10)
    ws_att.cell(r, 5).font = font(GREEN, True, 10)
    ws_att.cell(r, 21).number_format = "0%"

# Data validation for attendance
dv_status_att = DataValidation(type="list", formula1='"Present,Absent"')
ws_att.add_data_validation(dv_status_att); dv_status_att.add("I5:I200")

dv_extra = DataValidation(type="list", formula1='"Yes,No"')
ws_att.add_data_validation(dv_extra); dv_extra.add("J5:J200")

dv_mode = DataValidation(type="list", formula1='"Online,Offline,Hybrid"')
ws_att.add_data_validation(dv_mode); dv_mode.add("K5:K200")

# Conditional: Present = green, Absent = red
ws_att.conditional_formatting.add("I5:I200",
    FormulaRule(formula=['I5="Present"'], fill=fill(GREEN_DARK), font=font(WHITE, True)))
ws_att.conditional_formatting.add("I5:I200",
    FormulaRule(formula=['I5="Absent"'], fill=fill(RED_DARK), font=font(WHITE, True)))

# Attendance % color scale
ws_att.conditional_formatting.add("U5:U200",
    ColorScaleRule(start_type="num", start_value=0, start_color="FF4444",
                   mid_type="num", mid_value=0.6, mid_color="FFD700",
                   end_type="num", end_value=1, end_color="00FF88"))

# Per-course summary on right
ws_att.merge_cells("A210:N210")
ws_att["A210"].value = "📊 COURSE-WISE ATTENDANCE SUMMARY"
ws_att["A210"].fill = fill(BG_MID); ws_att["A210"].font = font(CYAN, True, 12); ws_att["A210"].alignment = align()

sum_headers = ["Course Code", "Course Title", "Total Classes", "Present", "Absent", "Extra", "Attendance %", "Status"]
sum_hcols = [GREEN, YELLOW, WHITE, GREEN, RED, ORANGE, GOLD, CYAN]
for col, (h, hc) in enumerate(zip(sum_headers, sum_hcols), 1):
    c = ws_att.cell(211, col, h)
    c.fill = fill(BG_CARD); c.font = font(hc, True, 10); c.alignment = align(); c.border = thin_border(CYAN)

unique_courses = [("CSE101", "Introduction to Programming"), ("MATH101", "Calculus I"),
                  ("PHY101", "Physics I"), ("CSE201", "Data Structures"), ("MATH201", "Calculus II")]

for i, (code, title) in enumerate(unique_courses):
    r = 212 + i
    bg = BG_DARK if i % 2 == 0 else BG_CARD
    ws_att.cell(r, 1, code).fill = fill(bg); ws_att.cell(r, 1).font = font(GREEN, True, 10)
    ws_att.cell(r, 2, title).fill = fill(bg); ws_att.cell(r, 2).font = font(YELLOW, False, 10)
    ws_att.cell(r, 3, f'=COUNTIF($E$5:$E$200,A{r})').fill = fill(bg); ws_att.cell(r, 3).font = font(WHITE, False, 10)
    ws_att.cell(r, 4, f'=COUNTIFS($E$5:$E$200,A{r},$I$5:$I$200,"Present")').fill = fill(bg); ws_att.cell(r, 4).font = font(GREEN, True, 10)
    ws_att.cell(r, 5, f'=COUNTIFS($E$5:$E$200,A{r},$I$5:$I$200,"Absent")').fill = fill(bg); ws_att.cell(r, 5).font = font(RED, True, 10)
    ws_att.cell(r, 6, f'=COUNTIFS($E$5:$E$200,A{r},$J$5:$J$200,"Yes")').fill = fill(bg); ws_att.cell(r, 6).font = font(ORANGE, False, 10)
    ws_att.cell(r, 7, f'=IF(C{r}=0,0,D{r}/C{r})').fill = fill(bg); ws_att.cell(r, 7).font = font(GOLD, True, 10)
    ws_att.cell(r, 7).number_format = "0.00%"
    ws_att.cell(r, 8, f'=IF(G{r}>=0.75,"✅ Good",IF(G{r}>=0.60,"⚠️ Warning","❌ Critical"))').fill = fill(bg)
    ws_att.cell(r, 8).font = font(WHITE, False, 10); ws_att.cell(r, 8).alignment = align()
    for col in range(1, 9):
        ws_att.cell(r, col).alignment = align()
        ws_att.cell(r, col).border = thin_border(GRAY_D)

# Grand total attendance for KPI
ws_att.cell(5, 19).value  # already set; set a named overall %
ws_att["S5"] = f'=IFERROR(COUNTIF(I5:I200,"Present")/COUNTA(I5:I200),0)'
ws_att["S5"].number_format = "0%"

# ══════════════════════════════════════════════════════════════
# SHEET 4: STUDY TRACKER
# ══════════════════════════════════════════════════════════════
ws_st = wb.create_sheet("📖 Study Tracker")
ws_st.sheet_view.showGridLines = False
ws_st.sheet_properties.tabColor = clr(YELLOW)

for row in ws_st.iter_rows(min_row=1, max_row=100, min_col=1, max_col=15):
    for cell in row:
        cell.fill = fill(BG_DARK)

merge_title(ws_st, "A1:O1", "📖  DONE STUDY TRACKER", BG_CARD, YELLOW, 16)
ws_st["A2"].value = "  Track which classes you have studied. Mark sessions as Done or Pending."
ws_st.merge_cells("A2:O2")
ws_st["A2"].fill = fill(BG_CARD); ws_st["A2"].font = font(GRAY_L, False, 10, italic=True); ws_st["A2"].alignment = align("left")

st_headers = ["#", "Semester", "Course Code", "Course Title", "Class Date", "Day",
              "Topic Covered", "Study Status", "Study Date", "Time Spent (hrs)",
              "Notes", "Difficulty", "Priority", "% Complete", "Remarks"]
st_colors = [GRAY_L, BLUE, GREEN, YELLOW, CYAN, PURPLE, WHITE, GOLD, TEAL, ORANGE, GRAY_L, RED, PINK, GREEN, CYAN]

ws_st.row_dimensions[4].height = 26
for col, (h, hc) in enumerate(zip(st_headers, st_colors), 1):
    c = ws_st.cell(4, col, h)
    c.fill = fill(BG_MID); c.font = font(hc, True, 10); c.alignment = align(); c.border = thin_border(CYAN)

st_widths = [4, 14, 12, 26, 12, 8, 24, 12, 12, 14, 18, 11, 10, 11, 18]
for i, w in enumerate(st_widths, 1):
    set_col_width(ws_st, i, w)

study_data = [
    ("1st Semester", "CSE101", "Introduction to Programming", "2026-01-05", "Variables & Data Types", "Done", "2026-01-06", 2.5, "Easy"),
    ("1st Semester", "MATH101", "Calculus I", "2026-01-05", "Limits", "Done", "2026-01-07", 3.0, "Medium"),
    ("1st Semester", "PHY101", "Physics I", "2026-01-05", "Motion", "Pending", "", 0, "Hard"),
    ("1st Semester", "CSE101", "Introduction to Programming", "2026-01-07", "Control Flow", "Done", "2026-01-08", 1.5, "Easy"),
    ("1st Semester", "MATH101", "Calculus I", "2026-01-07", "Derivatives", "Pending", "", 0, "Hard"),
    ("2nd Semester", "CSE201", "Data Structures", "2026-02-02", "Arrays", "Done", "2026-02-03", 2.0, "Medium"),
    ("2nd Semester", "CSE201", "Data Structures", "2026-02-04", "Linked Lists", "Pending", "", 0, "Hard"),
    ("2nd Semester", "MATH201", "Calculus II", "2026-02-04", "Integration", "Pending", "", 0, "Hard"),
]

for i, row_data in enumerate(study_data):
    r = 5 + i
    sem, code, title, cls_date, topic, status, study_date, hrs, diff = row_data
    bg = BG_DARK if i % 2 == 0 else BG_CARD
    ws_st.row_dimensions[r].height = 20

    vals = [i+1, sem, code, title, cls_date, datetime.strptime(cls_date, "%Y-%m-%d").strftime("%a"),
            topic, status, study_date if study_date else "", hrs if hrs else "",
            "", diff, "High" if diff == "Hard" else "Medium",
            f'=IF(H{r}="Done",1,0)',
            f'=IF(H{r}="Done","✅ Complete","⏳ Pending")']

    for col, val in enumerate(vals, 1):
        c = ws_st.cell(r, col, val)
        c.fill = fill(bg); c.font = font(WHITE, False, 10); c.alignment = align()
        c.border = thin_border(GRAY_D)

    ws_st.cell(r, 3).font = font(GREEN, True, 10)
    ws_st.cell(r, 5).font = font(CYAN, False, 10)

dv_study = DataValidation(type="list", formula1='"Done,Pending,In Progress"')
ws_st.add_data_validation(dv_study); dv_study.add("H5:H100")

dv_diff = DataValidation(type="list", formula1='"Easy,Medium,Hard"')
ws_st.add_data_validation(dv_diff); dv_diff.add("L5:L100")

dv_prio = DataValidation(type="list", formula1='"High,Medium,Low"')
ws_st.add_data_validation(dv_prio); dv_prio.add("M5:M100")

ws_st.conditional_formatting.add("H5:H100",
    FormulaRule(formula=['H5="Done"'], fill=fill(GREEN_DARK), font=font(WHITE, True)))
ws_st.conditional_formatting.add("H5:H100",
    FormulaRule(formula=['H5="Pending"'], fill=fill(BG_MID), font=font(YELLOW, True)))

# Summary at L5 (used by Dashboard KPI)
ws_st["L5"] = f'=IFERROR(COUNTIF(H5:H100,"Done")/COUNTA(H5:H100),0)'
ws_st["L5"].number_format = "0%"

# Course Progress Summary
r_sum = 5 + len(study_data) + 2
ws_st.merge_cells(f"A{r_sum}:O{r_sum}")
ws_st[f"A{r_sum}"].value = "📊 COURSE-WISE STUDY PROGRESS"
ws_st[f"A{r_sum}"].fill = fill(BG_MID); ws_st[f"A{r_sum}"].font = font(YELLOW, True, 12); ws_st[f"A{r_sum}"].alignment = align()

psum_headers = ["Course", "Total Sessions", "Done", "Pending", "Progress %", "Time Spent"]
for col, h in enumerate(psum_headers, 1):
    c = ws_st.cell(r_sum+1, col, h)
    c.fill = fill(BG_CARD); c.font = font(CYAN, True, 10); c.alignment = align(); c.border = thin_border(CYAN)

for i, (code, title) in enumerate([("CSE101", "Intro to Prog"), ("MATH101", "Calculus I"),
                                    ("PHY101", "Physics I"), ("CSE201", "Data Structures"), ("MATH201", "Calculus II")]):
    r = r_sum + 2 + i
    bg = BG_DARK if i % 2 == 0 else BG_CARD
    ws_st.cell(r, 1, code).fill = fill(bg); ws_st.cell(r, 1).font = font(GREEN, True, 10)
    ws_st.cell(r, 2, f'=COUNTIF($C$5:$C$100,A{r})').fill = fill(bg)
    ws_st.cell(r, 3, f'=COUNTIFS($C$5:$C$100,A{r},$H$5:$H$100,"Done")').fill = fill(bg); ws_st.cell(r, 3).font = font(GREEN, True, 10)
    ws_st.cell(r, 4, f'=COUNTIFS($C$5:$C$100,A{r},$H$5:$H$100,"Pending")').fill = fill(bg); ws_st.cell(r, 4).font = font(RED, True, 10)
    ws_st.cell(r, 5, f'=IFERROR(C{r}/B{r},0)').fill = fill(bg); ws_st.cell(r, 5).font = font(GOLD, True, 10)
    ws_st.cell(r, 5).number_format = "0%"
    ws_st.cell(r, 6, f'=SUMIF($C$5:$C$100,A{r},$J$5:$J$100)&" hrs"').fill = fill(bg); ws_st.cell(r, 6).font = font(ORANGE, False, 10)
    for col in range(1, 7):
        ws_st.cell(r, col).alignment = align(); ws_st.cell(r, col).border = thin_border(GRAY_D)
    for col in range(2, 6):
        ws_st.cell(r, col).fill = fill(bg); ws_st.cell(r, col).font = font(WHITE, False, 10)

# ══════════════════════════════════════════════════════════════
# SHEET 5: TASKS
# ══════════════════════════════════════════════════════════════
ws_task = wb.create_sheet("📝 Tasks")
ws_task.sheet_view.showGridLines = False
ws_task.sheet_properties.tabColor = clr(ORANGE)

for row in ws_task.iter_rows(min_row=1, max_row=220, min_col=1, max_col=14):
    for cell in row:
        cell.fill = fill(BG_DARK)

merge_title(ws_task, "A1:N1", "📝  DAILY TASK & ROUTINE MANAGER", BG_CARD, ORANGE, 16)
ws_task["A2"].value = "  Add, update and track all daily tasks with priority, deadline and status"
ws_task.merge_cells("A2:N2"); ws_task["A2"].fill = fill(BG_CARD)
ws_task["A2"].font = font(GRAY_L, False, 10, italic=True); ws_task["A2"].alignment = align("left")

task_headers = ["#", "Task Name", "Category", "Description", "Date", "Day",
                "Status", "Priority", "Start Time", "End Time", "Deadline",
                "Days Left", "Completed On", "Notes"]
task_hcols = [GRAY_L, YELLOW, CYAN, WHITE, BLUE, PURPLE, GOLD, RED, TEAL, TEAL, ORANGE, PINK, GREEN, GRAY_L]

ws_task.row_dimensions[4].height = 26
for col, (h, hc) in enumerate(zip(task_headers, task_hcols), 1):
    c = ws_task.cell(4, col, h)
    c.fill = fill(BG_MID); c.font = font(hc, True, 10); c.alignment = align(); c.border = thin_border(CYAN)

task_widths = [4, 26, 14, 30, 12, 8, 12, 10, 10, 10, 12, 10, 14, 20]
for i, w in enumerate(task_widths, 1):
    set_col_width(ws_task, i, w)

tasks_sample = [
    ("Complete Assignment 3", "Study", "Submit DSA assignment", "2026-05-22", "Pending", "High", "09:00 AM", "11:00 AM", "2026-05-22"),
    ("Buy Stationery", "Shopping", "Buy pens, notebook", "2026-05-20", "Completed", "Low", "04:00 PM", "05:00 PM", "2026-05-20"),
    ("Study Calculus", "Study", "Chapter 5 exercises", "2026-05-20", "Pending", "High", "08:00 PM", "10:00 PM", "2026-05-21"),
    ("Exercise", "Personal", "Morning workout", "2026-05-21", "Incomplete", "Medium", "06:00 AM", "07:00 AM", "2026-05-21"),
    ("Lab Report", "University", "Physics lab report", "2026-05-23", "Pending", "High", "02:00 PM", "04:00 PM", "2026-05-24"),
    ("Read Novel", "Personal", "Read 30 pages", "2026-05-20", "Completed", "Low", "09:00 PM", "10:00 PM", "2026-05-20"),
]

for i, t in enumerate(tasks_sample):
    r = 5 + i
    name, cat, desc, date_str, status, prio, start, end, deadline = t
    d = datetime.strptime(date_str, "%Y-%m-%d")
    bg = BG_DARK if i % 2 == 0 else BG_CARD
    ws_task.row_dimensions[r].height = 22

    vals = [i+1, name, cat, desc, date_str, d.strftime("%a"),
            status, prio, start, end, deadline,
            f'=IFERROR(DATEVALUE(K{r})-TODAY(),"—")',
            "" if status != "Completed" else date_str, ""]

    for col, val in enumerate(vals, 1):
        c = ws_task.cell(r, col, val)
        c.fill = fill(bg); c.font = font(WHITE, False, 10)
        c.alignment = align(); c.border = thin_border(GRAY_D)

    ws_task.cell(r, 2).font = font(YELLOW, False, 10)
    ws_task.cell(r, 3).font = font(CYAN, False, 10)
    ws_task.cell(r, 5).font = font(BLUE, False, 10)

dv_task_status = DataValidation(type="list", formula1='"Pending,Completed,Incomplete,In Progress"')
ws_task.add_data_validation(dv_task_status); dv_task_status.add("G5:G200")

dv_task_cat = DataValidation(type="list", formula1='"Study,Shopping,Personal,University,Assignment,Exam,Work,Other"')
ws_task.add_data_validation(dv_task_cat); dv_task_cat.add("C5:C200")

dv_task_prio = DataValidation(type="list", formula1='"High,Medium,Low"')
ws_task.add_data_validation(dv_task_prio); dv_task_prio.add("H5:H200")

ws_task.conditional_formatting.add("G5:G200",
    FormulaRule(formula=['G5="Completed"'], fill=fill(GREEN_DARK), font=font(WHITE, True)))
ws_task.conditional_formatting.add("G5:G200",
    FormulaRule(formula=['G5="Pending"'], fill=fill("4A3F00"), font=font(YELLOW, True)))
ws_task.conditional_formatting.add("G5:G200",
    FormulaRule(formula=['G5="Incomplete"'], fill=fill(RED_DARK), font=font(WHITE, True)))

ws_task.conditional_formatting.add("H5:H200",
    FormulaRule(formula=['H5="High"'], font=font(RED, True)))
ws_task.conditional_formatting.add("H5:H200",
    FormulaRule(formula=['H5="Low"'], font=font(GREEN, True)))

# Task summary
ws_task.merge_cells("A210:N210")
ws_task["A210"].value = "📊 TASK SUMMARY"
ws_task["A210"].fill = fill(BG_MID); ws_task["A210"].font = font(ORANGE, True, 12); ws_task["A210"].alignment = align()
labels = [("Pending", "G5:G200", YELLOW), ("Completed", "G5:G200", GREEN), ("Incomplete", "G5:G200", RED)]
for i, (lbl, rng, col) in enumerate(labels):
    ws_task.cell(211, i*2+1).value = lbl; ws_task.cell(211, i*2+1).font = font(col, True, 10); ws_task.cell(211, i*2+1).alignment = align()
    ws_task.cell(212, i*2+1).value = f'=COUNTIF({rng},"{lbl}")'
    ws_task.cell(212, i*2+1).font = font(WHITE, True, 14); ws_task.cell(212, i*2+1).alignment = align()
    ws_task.cell(211, i*2+1).fill = fill(BG_CARD); ws_task.cell(212, i*2+1).fill = fill(BG_CARD)

# ══════════════════════════════════════════════════════════════
# SHEET 6: EXPENSES
# ══════════════════════════════════════════════════════════════
ws_exp = wb.create_sheet("💸 Expenses")
ws_exp.sheet_view.showGridLines = False
ws_exp.sheet_properties.tabColor = clr(RED)

for row in ws_exp.iter_rows(min_row=1, max_row=220, min_col=1, max_col=12):
    for cell in row:
        cell.fill = fill(BG_DARK)

merge_title(ws_exp, "A1:L1", "💸  DAILY BUY & EXPENSE TRACKER", BG_CARD, RED, 16)
ws_exp["A2"].value = "  Track all purchases, expenses and payment status"
ws_exp.merge_cells("A2:L2"); ws_exp["A2"].fill = fill(BG_CARD)
ws_exp["A2"].font = font(GRAY_L, False, 10, italic=True); ws_exp["A2"].alignment = align("left")

exp_headers = ["#", "Date", "Day", "Item Name", "Amount (৳)", "Category", "Quantity",
               "Total (৳)", "Payment Status", "Payment Method", "Notes", "Month"]
exp_hcols = [GRAY_L, CYAN, PURPLE, YELLOW, RED, ORANGE, WHITE, GOLD, GREEN, BLUE, GRAY_L, TEAL]

ws_exp.row_dimensions[4].height = 26
for col, (h, hc) in enumerate(zip(exp_headers, exp_hcols), 1):
    c = ws_exp.cell(4, col, h)
    c.fill = fill(BG_MID); c.font = font(hc, True, 10); c.alignment = align(); c.border = thin_border(CYAN)

exp_widths = [4, 12, 8, 24, 12, 14, 10, 12, 14, 16, 20, 12]
for i, w in enumerate(exp_widths, 1):
    set_col_width(ws_exp, i, w)

exp_data = [
    ("2026-05-18", "Rice 5kg", 280, "Food", 1),
    ("2026-05-18", "Vegetables", 120, "Food", 1),
    ("2026-05-19", "Bus Fare", 40, "Transport", 2),
    ("2026-05-19", "Notebook", 60, "Stationery", 2),
    ("2026-05-20", "Lunch", 80, "Food", 1),
    ("2026-05-20", "Mobile Recharge", 100, "Personal", 1),
    ("2026-05-20", "Pen Set", 45, "Stationery", 1),
    ("2026-05-20", "Medicine", 150, "Health", 1),
]

payment_status = ["Received", "Received", "Received", "Pending", "Received", "Pending", "Received", "Pending"]
methods = ["Cash", "Cash", "Cash", "Cash", "bKash", "Nagad", "Cash", "Cash"]

for i, (exp, pstat, meth) in enumerate(zip(exp_data, payment_status, methods)):
    r = 5 + i
    date_str, item, price, cat, qty = exp
    d = datetime.strptime(date_str, "%Y-%m-%d")
    bg = BG_DARK if i % 2 == 0 else BG_CARD
    ws_exp.row_dimensions[r].height = 20

    vals = [i+1, date_str, d.strftime("%a"), item, price, cat, qty,
            f"=E{r}*G{r}", pstat, meth, "", f'=TEXT(B{r},"mmmm")']

    for col, val in enumerate(vals, 1):
        c = ws_exp.cell(r, col, val)
        c.fill = fill(bg); c.font = font(WHITE, False, 10)
        c.alignment = align(); c.border = thin_border(GRAY_D)

    ws_exp.cell(r, 2).font = font(CYAN, False, 10)
    ws_exp.cell(r, 4).font = font(YELLOW, False, 10)
    ws_exp.cell(r, 5).font = font(RED, True, 10)
    ws_exp.cell(r, 8).font = font(GOLD, True, 10)
    ws_exp.cell(r, 5).number_format = "#,##0"
    ws_exp.cell(r, 8).number_format = "#,##0"

dv_pay = DataValidation(type="list", formula1='"Pending,Received"')
ws_exp.add_data_validation(dv_pay); dv_pay.add("I5:I200")

dv_pmethod = DataValidation(type="list", formula1='"Cash,bKash,Nagad,Rocket,Bank Transfer,Card"')
ws_exp.add_data_validation(dv_pmethod); dv_pmethod.add("J5:J200")

dv_cat_exp = DataValidation(type="list", formula1='"Food,Transport,Stationery,Personal,University,Health,Shopping,Entertainment,Other"')
ws_exp.add_data_validation(dv_cat_exp); dv_cat_exp.add("F5:F200")

ws_exp.conditional_formatting.add("I5:I200",
    FormulaRule(formula=['I5="Received"'], fill=fill(GREEN_DARK), font=font(WHITE, True)))
ws_exp.conditional_formatting.add("I5:I200",
    FormulaRule(formula=['I5="Pending"'], fill=fill("4A3F00"), font=font(YELLOW, True)))

# Totals row
r_total = 5 + len(exp_data) + 1
ws_exp.merge_cells(f"A{r_total}:G{r_total}")
ws_exp[f"A{r_total}"].value = "TOTAL EXPENSE"
ws_exp[f"A{r_total}"].fill = fill(BG_MID); ws_exp[f"A{r_total}"].font = font(RED, True, 12); ws_exp[f"A{r_total}"].alignment = align("right")
ws_exp.cell(r_total, 8).value = f"=SUM(H5:H{r_total-2})"
ws_exp.cell(r_total, 8).fill = fill(BG_MID); ws_exp.cell(r_total, 8).font = font(GOLD, True, 13); ws_exp.cell(r_total, 8).alignment = align()
ws_exp.cell(r_total, 8).number_format = "#,##0"

# Category summary
r_cat = r_total + 2
ws_exp.merge_cells(f"A{r_cat}:L{r_cat}")
ws_exp[f"A{r_cat}"].value = "📊 CATEGORY-WISE EXPENSE SUMMARY"
ws_exp[f"A{r_cat}"].fill = fill(BG_MID); ws_exp[f"A{r_cat}"].font = font(ORANGE, True, 12); ws_exp[f"A{r_cat}"].alignment = align()

cats = ["Food", "Transport", "Stationery", "Personal", "University", "Health", "Shopping", "Entertainment", "Other"]
for i, cat in enumerate(cats):
    r = r_cat + 1 + i
    bg = BG_DARK if i % 2 == 0 else BG_CARD
    ws_exp.cell(r, 1, cat).fill = fill(bg); ws_exp.cell(r, 1).font = font(ORANGE, True, 10); ws_exp.cell(r, 1).alignment = align()
    ws_exp.cell(r, 2, f'=SUMIF($F$5:$F$200,A{r},$H$5:$H$200)').fill = fill(bg)
    ws_exp.cell(r, 2).font = font(GOLD, True, 11); ws_exp.cell(r, 2).alignment = align()
    ws_exp.cell(r, 2).number_format = "#,##0"
    ws_exp.cell(r, 2).border = thin_border(GRAY_D); ws_exp.cell(r, 1).border = thin_border(GRAY_D)

# ══════════════════════════════════════════════════════════════
# SHEET 7: MONEY MANAGER
# ══════════════════════════════════════════════════════════════
ws_mm = wb.create_sheet("💰 Money Manager")
ws_mm.sheet_view.showGridLines = False
ws_mm.sheet_properties.tabColor = clr(GREEN)

for row in ws_mm.iter_rows(min_row=1, max_row=120, min_col=1, max_col=16):
    for cell in row:
        cell.fill = fill(BG_DARK)

merge_title(ws_mm, "A1:P1", "💰  FUND & MONEY MANAGER", BG_CARD, GREEN, 16)
ws_mm["A2"].value = "  Manage all your accounts: Bank, bKash, Nagad, Rocket, Cash"
ws_mm.merge_cells("A2:P2"); ws_mm["A2"].fill = fill(BG_CARD)
ws_mm["A2"].font = font(GRAY_L, False, 10, italic=True); ws_mm["A2"].alignment = align("left")

# Accounts summary
acc_headers = ["Account Name", "Account Type", "Account No", "Opening Balance (৳)", "Total Deposit (৳)",
               "Total Withdraw (৳)", "Current Balance (৳)", "Status"]
acc_hcols = [GREEN, CYAN, BLUE, GOLD, GREEN, RED, YELLOW, TEAL]

ws_mm.row_dimensions[4].height = 26
for col, (h, hc) in enumerate(zip(acc_headers, acc_hcols), 1):
    c = ws_mm.cell(4, col, h)
    c.fill = fill(BG_MID); c.font = font(hc, True, 10); c.alignment = align(); c.border = thin_border(CYAN)

accounts = [
    ("My bKash", "bKash", "01XXXXXXXXX", 500),
    ("My Bank", "Bank", "ACC-XXXXXXX", 5000),
    ("Nagad", "Nagad", "01XXXXXXXXX", 200),
    ("Hand Cash", "Cash", "—", 1000),
]

for i, (name, acc_type, acc_no, opening) in enumerate(accounts):
    r = 5 + i
    bg = BG_DARK if i % 2 == 0 else BG_CARD
    ws_mm.row_dimensions[r].height = 24
    ws_mm.cell(r, 1, name).fill = fill(bg); ws_mm.cell(r, 1).font = font(GREEN, True, 11)
    ws_mm.cell(r, 2, acc_type).fill = fill(bg); ws_mm.cell(r, 2).font = font(CYAN, True, 10)
    ws_mm.cell(r, 3, acc_no).fill = fill(bg); ws_mm.cell(r, 3).font = font(GRAY_L, False, 10)
    ws_mm.cell(r, 4, opening).fill = fill(bg); ws_mm.cell(r, 4).font = font(GOLD, True, 11); ws_mm.cell(r, 4).number_format = "#,##0"
    ws_mm.cell(r, 5, f'=SUMIFS($L$25:$L$200,$J$25:$J$200,A{r},$K$25:$K$200,"Deposit")+D{r}').fill = fill(bg)
    ws_mm.cell(r, 5).font = font(GREEN, True, 11); ws_mm.cell(r, 5).number_format = "#,##0"
    ws_mm.cell(r, 6, f'=SUMIFS($L$25:$L$200,$J$25:$J$200,A{r},$K$25:$K$200,"Withdraw")').fill = fill(bg)
    ws_mm.cell(r, 6).font = font(RED, True, 11); ws_mm.cell(r, 6).number_format = "#,##0"
    ws_mm.cell(r, 7, f'=E{r}-F{r}').fill = fill(bg); ws_mm.cell(r, 7).font = font(YELLOW, True, 13); ws_mm.cell(r, 7).number_format = "#,##0"
    ws_mm.cell(r, 8, "Active").fill = fill(bg); ws_mm.cell(r, 8).font = font(GREEN, False, 10); ws_mm.cell(r, 8).alignment = align()
    for col in range(1, 9):
        ws_mm.cell(r, col).alignment = align(); ws_mm.cell(r, col).border = thin_border(GRAY_D)

# Account status validation
dv_acc_status = DataValidation(type="list", formula1='"Active,Closed"')
ws_mm.add_data_validation(dv_acc_status); dv_acc_status.add("H5:H20")

# Account type validation
dv_acc_type = DataValidation(type="list", formula1='"bKash,Bank,Nagad,Rocket,Cash,Other"')
ws_mm.add_data_validation(dv_acc_type); dv_acc_type.add("B5:B20")

# Total balance
r_bal = 5 + len(accounts) + 1
ws_mm.merge_cells(f"A{r_bal}:F{r_bal}")
ws_mm[f"A{r_bal}"].value = "💰 TOTAL BALANCE (ALL ACCOUNTS)"
ws_mm[f"A{r_bal}"].fill = fill(BG_MID); ws_mm[f"A{r_bal}"].font = font(GOLD, True, 12); ws_mm[f"A{r_bal}"].alignment = align("right")
ws_mm.cell(r_bal, 7).value = f"=SUM(G5:G{r_bal-2})"
ws_mm.cell(r_bal, 7).fill = fill(BG_MID); ws_mm.cell(r_bal, 7).font = font(GREEN, True, 16); ws_mm.cell(r_bal, 7).alignment = align()
ws_mm.cell(r_bal, 7).number_format = "#,##0"

# Transaction log
r_txn = r_bal + 3
ws_mm.merge_cells(f"A{r_txn}:P{r_txn}")
ws_mm[f"A{r_txn}"].value = "📋 TRANSACTION HISTORY"
ws_mm[f"A{r_txn}"].fill = fill(BG_MID); ws_mm[f"A{r_txn}"].font = font(CYAN, True, 13); ws_mm[f"A{r_txn}"].alignment = align()

txn_headers = ["#", "Date", "Day", "Time", "Description", "Category",
               "Account", "Type", "Amount (৳)", "Balance After (৳)", "Ref No", "Notes"]
txn_hcols = [GRAY_L, CYAN, PURPLE, TEAL, YELLOW, ORANGE, GREEN, BLUE, GOLD, GREEN, GRAY_L, GRAY_L]
txn_row = r_txn + 1

for col, (h, hc) in enumerate(zip(txn_headers, txn_hcols), 1):
    c = ws_mm.cell(txn_row, col, h)
    c.fill = fill(BG_CARD); c.font = font(hc, True, 10); c.alignment = align(); c.border = thin_border(CYAN)

# Rename columns J and K for SUMIFS formula references
# J = Account name, K = Type (Deposit/Withdraw), L = Amount
# Map: col7=Account(J col10), col8=Type(K col11), col9=Amount(L col12) in 0-indexed...
# Actually in txn table: col1=#, col2=date, col3=day, col4=time, col5=desc, col6=cat
# col7=account, col8=type, col9=amount, col10=balance after, col11=ref, col12=notes
# J in Excel = col10, K=col11, L=col12. Our formula references J/K/L columns (10/11/12).
# Rewrite sumifs to match actual col positions: col7=Account(G=col7), col8=Type(H=col8), col9=Amount(I=col9)

# Fix account balance formulas to match actual columns
for i, (name, acc_type, acc_no, opening) in enumerate(accounts):
    r = 5 + i
    data_start = txn_row + 1
    ws_mm.cell(r, 5).value = f'=SUMIFS($I${data_start}:$I$200,$G${data_start}:$G$200,A{r},$H${data_start}:$H$200,"Deposit")+D{r}'
    ws_mm.cell(r, 6).value = f'=SUMIFS($I${data_start}:$I$200,$G${data_start}:$G$200,A{r},$H${data_start}:$H$200,"Withdraw")'

txn_data = [
    ("2026-05-18", "10:30 AM", "Received from home", "Transfer", "My bKash", "Deposit", 2000),
    ("2026-05-18", "11:00 AM", "Bus fare", "Transport", "Hand Cash", "Withdraw", 40),
    ("2026-05-19", "09:00 AM", "Rice purchase", "Food", "Hand Cash", "Withdraw", 280),
    ("2026-05-19", "03:00 PM", "Mobile bill", "Personal", "My bKash", "Withdraw", 100),
    ("2026-05-20", "08:00 AM", "Pocket money", "Transfer", "Hand Cash", "Deposit", 500),
]

for i, txn in enumerate(txn_data):
    r = txn_row + 1 + i
    date_str, time_s, desc, cat, acc, txn_type, amt = txn
    d = datetime.strptime(date_str, "%Y-%m-%d")
    bg = BG_DARK if i % 2 == 0 else BG_CARD
    ws_mm.row_dimensions[r].height = 20

    vals = [i+1, date_str, d.strftime("%a"), time_s, desc, cat, acc, txn_type, amt, "", f"TXN-{1000+i}", ""]
    for col, val in enumerate(vals, 1):
        c = ws_mm.cell(r, col, val)
        c.fill = fill(bg); c.font = font(WHITE, False, 10)
        c.alignment = align(); c.border = thin_border(GRAY_D)

    ws_mm.cell(r, 9).number_format = "#,##0"
    ws_mm.cell(r, 9).font = font(GOLD, True, 10) if txn_type == "Deposit" else font(RED, True, 10)

dv_txn_type = DataValidation(type="list", formula1='"Deposit,Withdraw"')
ws_mm.add_data_validation(dv_txn_type); dv_txn_type.add(f"H{txn_row+1}:H200")

dv_txn_acc = DataValidation(type="list", formula1='"My bKash,My Bank,Nagad,Hand Cash"')
ws_mm.add_data_validation(dv_txn_acc); dv_txn_acc.add(f"G{txn_row+1}:G200")

ws_mm.conditional_formatting.add(f"H{txn_row+1}:H200",
    FormulaRule(formula=[f'H{txn_row+1}="Deposit"'], fill=fill(GREEN_DARK), font=font(WHITE, True)))
ws_mm.conditional_formatting.add(f"H{txn_row+1}:H200",
    FormulaRule(formula=[f'H{txn_row+1}="Withdraw"'], fill=fill(RED_DARK), font=font(WHITE, True)))

mm_widths = [4, 12, 8, 10, 26, 14, 16, 10, 14, 14, 12, 20]
for i, w in enumerate(mm_widths, 1):
    set_col_width(ws_mm, i, w)

# ══════════════════════════════════════════════════════════════
# SHEET 8: CGPA
# ══════════════════════════════════════════════════════════════
ws_cgpa = wb.create_sheet("🏆 CGPA")
ws_cgpa.sheet_view.showGridLines = False
ws_cgpa.sheet_properties.tabColor = clr(PURPLE)

for row in ws_cgpa.iter_rows(min_row=1, max_row=80, min_col=1, max_col=14):
    for cell in row:
        cell.fill = fill(BG_DARK)

merge_title(ws_cgpa, "A1:N1", "🏆  CGPA & RESULT TRACKER", BG_CARD, PURPLE, 16)
ws_cgpa["A2"].value = "  Enter marks per course, GPA auto-calculated. CGPA updates automatically."
ws_cgpa.merge_cells("A2:N2"); ws_cgpa["A2"].fill = fill(BG_CARD)
ws_cgpa["A2"].font = font(GRAY_L, False, 10, italic=True); ws_cgpa["A2"].alignment = align("left")

cgpa_headers = ["Semester", "Course Code", "Course Title", "Credit", "Mid (30)", "Final (60)",
                "Assignment (5)", "Attendance (5)", "Total (100)", "GPA (4.0)", "Grade",
                "Grade Point", "Weighted Points", "Remarks"]
cgpa_hcols = [BLUE, GREEN, YELLOW, ORANGE, CYAN, CYAN, PINK, TEAL, GOLD, PURPLE, RED, RED, PURPLE, GRAY_L]

ws_cgpa.row_dimensions[4].height = 26
for col, (h, hc) in enumerate(zip(cgpa_headers, cgpa_hcols), 1):
    c = ws_cgpa.cell(4, col, h)
    c.fill = fill(BG_MID); c.font = font(hc, True, 9); c.alignment = align(); c.border = thin_border(CYAN)

cgpa_widths = [14, 12, 26, 8, 10, 10, 13, 13, 11, 10, 8, 12, 15, 16]
for i, w in enumerate(cgpa_widths, 1):
    set_col_width(ws_cgpa, i, w)

cgpa_data = [
    ("1st Semester", "CSE101", "Introduction to Programming", 3, 26, 52, 4, 5),
    ("1st Semester", "MATH101", "Calculus I", 3, 22, 48, 3, 5),
    ("1st Semester", "PHY101", "Physics I", 3, 20, 40, 4, 4),
    ("1st Semester", "ENG101", "English Communication", 2, 28, 55, 5, 5),
    ("2nd Semester", "CSE201", "Data Structures", 3, 24, 50, 4, 5),
    ("2nd Semester", "MATH201", "Calculus II", 3, 19, 42, 3, 4),
]

for i, row_data in enumerate(cgpa_data):
    r = 5 + i
    sem, code, title, credit, mid, final, asgn, att = row_data
    bg = BG_DARK if i % 2 == 0 else BG_CARD
    ws_cgpa.row_dimensions[r].height = 22

    total_formula = f"=E{r}+F{r}+G{r}+H{r}"
    grade_formula = f'=IF(I{r}>=80,"A+",IF(I{r}>=75,"A",IF(I{r}>=70,"A-",IF(I{r}>=65,"B+",IF(I{r}>=60,"B",IF(I{r}>=55,"B-",IF(I{r}>=50,"C+",IF(I{r}>=45,"C",IF(I{r}>=40,"D","F")))))))))'
    gpa_formula = f'=IF(I{r}>=80,4.0,IF(I{r}>=75,3.75,IF(I{r}>=70,3.5,IF(I{r}>=65,3.25,IF(I{r}>=60,3.0,IF(I{r}>=55,2.75,IF(I{r}>=50,2.5,IF(I{r}>=45,2.25,IF(I{r}>=40,2.0,0.0)))))))))'
    weighted_formula = f"=D{r}*J{r}"

    vals = [sem, code, title, credit, mid, final, asgn, att,
            total_formula, gpa_formula, grade_formula,
            f"=J{r}", weighted_formula, ""]

    for col, val in enumerate(vals, 1):
        c = ws_cgpa.cell(r, col, val)
        c.fill = fill(bg); c.font = font(WHITE, False, 10)
        c.alignment = align(); c.border = thin_border(GRAY_D)

    ws_cgpa.cell(r, 2).font = font(GREEN, True, 10)
    ws_cgpa.cell(r, 9).font = font(GOLD, True, 11)
    ws_cgpa.cell(r, 10).font = font(PURPLE, True, 11)

# Marks validation
dv_marks_cgpa = DataValidation(type="decimal", operator="between", formula1="0", formula2="100")
ws_cgpa.add_data_validation(dv_marks_cgpa)
ws_cgpa.add_data_validation(dv_marks_cgpa)
dv_marks_cgpa.add("E5:H50")

# Conditional formatting grades
grade_rules = [("A+", GREEN_DARK), ("A", GREEN_MID), ("A-", "007744"), ("B+", "0055AA"), ("B", BLUE), ("F", RED_DARK)]
for grade, bg_col in grade_rules:
    ws_cgpa.conditional_formatting.add("K5:K50",
        FormulaRule(formula=[f'K5="{grade}"'], fill=fill(bg_col), font=font(WHITE, True)))

# Semester SGPA summaries
r_sgpa = 5 + len(cgpa_data) + 2
ws_cgpa.merge_cells(f"A{r_sgpa}:N{r_sgpa}")
ws_cgpa[f"A{r_sgpa}"].value = "📊 SEMESTER WISE SGPA & CGPA"
ws_cgpa[f"A{r_sgpa}"].fill = fill(BG_MID); ws_cgpa[f"A{r_sgpa}"].font = font(PURPLE, True, 13); ws_cgpa[f"A{r_sgpa}"].alignment = align()

sgpa_headers = ["Semester", "Total Credits", "Total Weighted Points", "SGPA", "Running CGPA", "Performance"]
for col, h in enumerate(sgpa_headers, 1):
    c = ws_cgpa.cell(r_sgpa+1, col, h)
    c.fill = fill(BG_CARD); c.font = font(CYAN, True, 10); c.alignment = align(); c.border = thin_border(CYAN)

semesters_for_sgpa = ["1st Semester", "2nd Semester"]
for i, sem in enumerate(semesters_for_sgpa):
    r = r_sgpa + 2 + i
    bg = BG_DARK if i % 2 == 0 else BG_CARD
    ws_cgpa.cell(r, 1, sem).fill = fill(bg); ws_cgpa.cell(r, 1).font = font(BLUE, True, 10)
    ws_cgpa.cell(r, 2, f'=SUMIF($A$5:$A$50,A{r},$D$5:$D$50)').fill = fill(bg)
    ws_cgpa.cell(r, 3, f'=SUMIF($A$5:$A$50,A{r},$M$5:$M$50)').fill = fill(bg)
    ws_cgpa.cell(r, 4, f'=IFERROR(C{r}/B{r},0)').fill = fill(bg); ws_cgpa.cell(r, 4).font = font(PURPLE, True, 12)
    ws_cgpa.cell(r, 4).number_format = "0.00"
    if i == 0:
        ws_cgpa.cell(r, 5, f'=D{r}').fill = fill(bg)
    else:
        prev_r = r - 1
        ws_cgpa.cell(r, 5, f'=IFERROR(SUM($C${r_sgpa+2}:C{r})/SUM($B${r_sgpa+2}:B{r}),0)').fill = fill(bg)
    ws_cgpa.cell(r, 5).font = font(GOLD, True, 13); ws_cgpa.cell(r, 5).number_format = "0.00"
    ws_cgpa.cell(r, 6, f'=IF(E{r}>=3.75,"🌟 Excellent",IF(E{r}>=3.5,"⭐ Very Good",IF(E{r}>=3.0,"✅ Good",IF(E{r}>=2.5,"📘 Average","⚠️ Below Avg"))))').fill = fill(bg)
    ws_cgpa.cell(r, 6).font = font(WHITE, False, 10)
    for col in range(1, 7):
        ws_cgpa.cell(r, col).alignment = align(); ws_cgpa.cell(r, col).border = thin_border(GRAY_D)
        if col not in [1, 4, 5, 6]:
            ws_cgpa.cell(r, col).fill = fill(bg); ws_cgpa.cell(r, col).font = font(WHITE, False, 10)

# Overall CGPA at H5 (dashboard reference)
ws_cgpa["H5"] = f'=IFERROR(SUM($M$5:$M$50)/SUM($D$5:$D$50),0)'
ws_cgpa["H5"].fill = fill(BG_DARK); ws_cgpa["H5"].font = font(GOLD, True, 14); ws_cgpa["H5"].alignment = align()
ws_cgpa["H5"].number_format = "0.00"

# ══════════════════════════════════════════════════════════════
# SHEET 9: CALENDAR VIEW
# ══════════════════════════════════════════════════════════════
ws_cal = wb.create_sheet("📅 Calendar")
ws_cal.sheet_view.showGridLines = False
ws_cal.sheet_properties.tabColor = clr(PINK)

for row in ws_cal.iter_rows(min_row=1, max_row=50, min_col=1, max_col=35):
    for cell in row:
        cell.fill = fill(BG_DARK)

merge_title(ws_cal, "A1:AI1", "📅  ATTENDANCE CALENDAR — MAY 2026", BG_CARD, PINK, 16)

day_names = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
month_start_day = 4  # May 2026 starts on Friday (index 4 in 0=Sun)
days_in_month = 31

for d, day in enumerate(day_names):
    c = ws_cal.cell(3, d+1, day)
    c.fill = fill(BG_MID); c.font = font(CYAN, True, 11); c.alignment = align()
    set_col_width(ws_cal, d+1, 14)

calendar_grid = []
week = [None] * month_start_day
for day in range(1, days_in_month + 1):
    week.append(day)
    if len(week) == 7:
        calendar_grid.append(week)
        week = []
while len(week) < 7:
    week.append(None)
if any(w is not None for w in week):
    calendar_grid.append(week)

# Attendance data for May 2026
att_may = {
    5: [("CSE101", "P"), ("MATH101", "P")],
    7: [("CSE101", "P"), ("PHY101", "A")],
    12: [("MATH101", "P"), ("CSE201", "P")],
    14: [("CSE101", "A"), ("CSE201", "P")],
    19: [("MATH101", "P")],
    20: [("CSE101", "P"), ("PHY101", "P")],
    21: [("CSE201", "A")],
    26: [("CSE101", "P"), ("MATH101", "P"), ("PHY101", "P")],
    28: [("CSE201", "P")],
}

for week_i, week_data in enumerate(calendar_grid):
    base_row = 4 + week_i * 5
    ws_cal.row_dimensions[base_row].height = 16
    ws_cal.row_dimensions[base_row+1].height = 14
    ws_cal.row_dimensions[base_row+2].height = 14
    ws_cal.row_dimensions[base_row+3].height = 14
    ws_cal.row_dimensions[base_row+4].height = 4

    for col_i, day_num in enumerate(week_data):
        col = col_i + 1
        if day_num is not None:
            is_today = (day_num == 20)
            day_bg = CYAN_DARK if is_today else BG_CARD
            c = ws_cal.cell(base_row, col, day_num)
            c.fill = fill(day_bg)
            c.font = font(WHITE if not is_today else BG_DARK, True, 12)
            c.alignment = align()
            c.border = thin_border(CYAN if is_today else GRAY_D)

            classes = att_may.get(day_num, [])
            for cls_i, (course, status) in enumerate(classes[:3]):
                c2 = ws_cal.cell(base_row + 1 + cls_i, col)
                c2.value = course
                c2.fill = fill(GREEN_DARK if status == "P" else RED_DARK)
                c2.font = font(WHITE, True, 8)
                c2.alignment = align()
        else:
            for rr in range(5):
                ws_cal.cell(base_row + rr, col).fill = fill(BG_DARK)

# Legend
legend_row = 4 + len(calendar_grid) * 5 + 1
ws_cal.merge_cells(f"A{legend_row}:G{legend_row}")
ws_cal[f"A{legend_row}"].value = "🟢 Green = Present   🔴 Red = Absent   🔵 Blue highlight = Today"
ws_cal[f"A{legend_row}"].fill = fill(BG_CARD); ws_cal[f"A{legend_row}"].font = font(GRAY_L, False, 10)
ws_cal[f"A{legend_row}"].alignment = align("left")

# ══════════════════════════════════════════════════════════════
# SHEET 10: FEE MANAGER
# ══════════════════════════════════════════════════════════════
ws_fee = wb.create_sheet("📑 Fee Manager")
ws_fee.sheet_view.showGridLines = False
ws_fee.sheet_properties.tabColor = clr(GOLD)

for row in ws_fee.iter_rows(min_row=1, max_row=80, min_col=1, max_col=12):
    for cell in row:
        cell.fill = fill(BG_DARK)

merge_title(ws_fee, "A1:L1", "📑  REGISTRATION FEE MANAGER", BG_CARD, GOLD, 16)
ws_fee["A2"].value = "  Track all semester registration fees, payment status and due amounts"
ws_fee.merge_cells("A2:L2"); ws_fee["A2"].fill = fill(BG_CARD)
ws_fee["A2"].font = font(GRAY_L, False, 10, italic=True); ws_fee["A2"].alignment = align("left")

fee_headers = ["#", "Semester", "Fee Type", "Amount (৳)", "Due Date", "Payment Date",
               "Payment Status", "Payment Method", "Receipt No", "Bank/Account", "Notes", "Penalty"]
fee_hcols = [GRAY_L, BLUE, CYAN, GOLD, RED, GREEN, YELLOW, TEAL, PURPLE, ORANGE, GRAY_L, RED]

ws_fee.row_dimensions[4].height = 26
for col, (h, hc) in enumerate(zip(fee_headers, fee_hcols), 1):
    c = ws_fee.cell(4, col, h)
    c.fill = fill(BG_MID); c.font = font(hc, True, 10); c.alignment = align(); c.border = thin_border(CYAN)

fee_widths = [4, 14, 20, 13, 12, 14, 14, 16, 13, 16, 20, 10]
for i, w in enumerate(fee_widths, 1):
    set_col_width(ws_fee, i, w)

fee_data = [
    ("1st Semester", "Hall Fee", 2500, "2024-02-01", "2024-01-28", "Paid"),
    ("1st Semester", "Faculty Fee", 3000, "2024-02-01", "2024-01-28", "Paid"),
    ("1st Semester", "Library Fee", 500, "2024-02-01", "2024-01-28", "Paid"),
    ("2nd Semester", "Hall Fee", 2500, "2024-08-01", "2024-07-30", "Paid"),
    ("2nd Semester", "Faculty Fee", 3000, "2024-08-01", "2024-07-30", "Paid"),
    ("3rd Semester", "Hall Fee", 2500, "2025-02-01", "", "Pending"),
    ("3rd Semester", "Faculty Fee", 3200, "2025-02-01", "", "Pending"),
]

for i, fd in enumerate(fee_data):
    r = 5 + i
    sem, fee_type, amt, due, paid_date, status = fd
    bg = BG_DARK if i % 2 == 0 else BG_CARD
    ws_fee.row_dimensions[r].height = 20

    vals = [i+1, sem, fee_type, amt, due, paid_date if paid_date else "",
            status, "Bank" if paid_date else "", f"RCP-{1000+i}" if paid_date else "", "", "", 0]

    for col, val in enumerate(vals, 1):
        c = ws_fee.cell(r, col, val)
        c.fill = fill(bg); c.font = font(WHITE, False, 10)
        c.alignment = align(); c.border = thin_border(GRAY_D)

    ws_fee.cell(r, 4).font = font(GOLD, True, 11); ws_fee.cell(r, 4).number_format = "#,##0"

dv_fee_status = DataValidation(type="list", formula1='"Paid,Pending,Overdue,Partial"')
ws_fee.add_data_validation(dv_fee_status); dv_fee_status.add("G5:G80")

dv_fee_sem = DataValidation(type="list", formula1='"1st Semester,2nd Semester,3rd Semester,4th Semester,5th Semester,6th Semester,7th Semester,8th Semester"')
ws_fee.add_data_validation(dv_fee_sem); dv_fee_sem.add("B5:B80")

ws_fee.conditional_formatting.add("G5:G80",
    FormulaRule(formula=['G5="Paid"'], fill=fill(GREEN_DARK), font=font(WHITE, True)))
ws_fee.conditional_formatting.add("G5:G80",
    FormulaRule(formula=['G5="Pending"'], fill=fill("4A3F00"), font=font(YELLOW, True)))
ws_fee.conditional_formatting.add("G5:G80",
    FormulaRule(formula=['G5="Overdue"'], fill=fill(RED_DARK), font=font(WHITE, True)))

r_ftotal = 5 + len(fee_data) + 1
ws_fee.merge_cells(f"A{r_ftotal}:C{r_ftotal}")
ws_fee[f"A{r_ftotal}"].value = "TOTAL FEES"
ws_fee[f"A{r_ftotal}"].fill = fill(BG_MID); ws_fee[f"A{r_ftotal}"].font = font(GOLD, True, 12); ws_fee[f"A{r_ftotal}"].alignment = align("right")
ws_fee.cell(r_ftotal, 4).value = f"=SUM(D5:D{r_ftotal-2})"
ws_fee.cell(r_ftotal, 4).fill = fill(BG_MID); ws_fee.cell(r_ftotal, 4).font = font(GOLD, True, 14); ws_fee.cell(r_ftotal, 4).alignment = align()
ws_fee.cell(r_ftotal, 4).number_format = "#,##0"

ws_fee.merge_cells(f"A{r_ftotal+1}:C{r_ftotal+1}")
ws_fee[f"A{r_ftotal+1}"].value = "PAID"
ws_fee[f"A{r_ftotal+1}"].fill = fill(BG_CARD); ws_fee[f"A{r_ftotal+1}"].font = font(GREEN, True, 11); ws_fee[f"A{r_ftotal+1}"].alignment = align("right")
ws_fee.cell(r_ftotal+1, 4).value = f'=SUMIF(G5:G{r_ftotal-2},"Paid",D5:D{r_ftotal-2})'
ws_fee.cell(r_ftotal+1, 4).fill = fill(BG_CARD); ws_fee.cell(r_ftotal+1, 4).font = font(GREEN, True, 13); ws_fee.cell(r_ftotal+1, 4).alignment = align()
ws_fee.cell(r_ftotal+1, 4).number_format = "#,##0"

ws_fee.merge_cells(f"A{r_ftotal+2}:C{r_ftotal+2}")
ws_fee[f"A{r_ftotal+2}"].value = "DUE / PENDING"
ws_fee[f"A{r_ftotal+2}"].fill = fill(BG_CARD); ws_fee[f"A{r_ftotal+2}"].font = font(RED, True, 11); ws_fee[f"A{r_ftotal+2}"].alignment = align("right")
ws_fee.cell(r_ftotal+2, 4).value = f'=SUMIF(G5:G{r_ftotal-2},"Pending",D5:D{r_ftotal-2})'
ws_fee.cell(r_ftotal+2, 4).fill = fill(BG_CARD); ws_fee.cell(r_ftotal+2, 4).font = font(RED, True, 13); ws_fee.cell(r_ftotal+2, 4).alignment = align()
ws_fee.cell(r_ftotal+2, 4).number_format = "#,##0"

# ══════════════════════════════════════════════════════════════
# SHEET 11: MY DETAILS (Student Info)
# ══════════════════════════════════════════════════════════════
ws_me = wb.create_sheet("👤 My Details")
ws_me.sheet_view.showGridLines = False
ws_me.sheet_properties.tabColor = clr(TEAL)

for row in ws_me.iter_rows(min_row=1, max_row=40, min_col=1, max_col=10):
    for cell in row:
        cell.fill = fill(BG_DARK)

merge_title(ws_me, "A1:J1", "👤  STUDENT PROFILE", BG_CARD, TEAL, 16)

profile_fields = [
    ("Full Name", "SHOVON", CYAN),
    ("Student ID", "20XXXXXXX", YELLOW),
    ("Registration No", "2020-XXXXX", ORANGE),
    ("Session", "2020-21", GREEN),
    ("Department", "CSE", PURPLE),
    ("Faculty", "Engineering", BLUE),
    ("University", "My University", TEAL),
    ("Semester", "3rd Semester", GOLD),
    ("Email", "shovon@email.com", PINK),
    ("Phone", "+880-XXXXXXXXX", WHITE),
    ("Hall/Hostel", "Hall Name", ORANGE),
    ("Advisor", "Dr. Someone", CYAN),
    ("Date of Birth", "", GRAY_L),
    ("Blood Group", "B+", RED),
    ("Address", "", GRAY_L),
]

for i, (field, val, col) in enumerate(profile_fields):
    r = 3 + i
    ws_me.row_dimensions[r].height = 24
    ws_me.merge_cells(f"A{r}:C{r}")
    c1 = ws_me.cell(r, 1, f"  {field}:")
    c1.fill = fill(BG_CARD); c1.font = font(GRAY_L, True, 11); c1.alignment = align("left")
    ws_me.merge_cells(f"D{r}:J{r}")
    c2 = ws_me.cell(r, 4, val)
    c2.fill = fill(BG_MID); c2.font = font(col, True if val else False, 12)
    c2.alignment = align("left"); c2.border = thin_border(GRAY_D)

for i in range(1, 11):
    set_col_width(ws_me, i, 12)

# ══════════════════════════════════════════════════════════════
# ROW HEIGHTS FOR DASHBOARD
# ══════════════════════════════════════════════════════════════
ws_dash.row_dimensions[1].height = 36
ws_dash.row_dimensions[2].height = 22
ws_dash.row_dimensions[3].height = 18

# ══════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════
out_path = "SHOVON_Academic_Manager.xlsx"
wb.save(out_path)
print("Saved:", out_path)
